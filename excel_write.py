import os
import time
from typing import Any, NoReturn, List

from constants import MONTHS

import openpyxl as xl


def delete_first_sheet(workbook: Any) -> NoReturn:
    first_sheet = workbook.get_sheet_names()[0]
    del workbook[first_sheet]


def write_headers(worksheet_object: Any) -> NoReturn:
    headers_tuple = (
        "Время суток",
        "Температура",
        "Давление",
        "Влажность",
        "Погодное явление",
        "Среднесуточная температура",
        "Магнитное поле",
        "Изменение давления",
    )
    for col_n in range(1, len(headers_tuple) + 1):
        worksheet_object.cell(row=1, column=col_n).value = headers_tuple[col_n - 1]


def write_to_excel(city_name: str, result_list: List) -> NoReturn:
    first_date = result_list[0][0]
    month_number = MONTHS[result_list[0][1]]
    file_name = f"excel_files/{city_name}-{first_date}-{month_number}.xlsx"
    wb = xl.Workbook()
    for day in result_list:
        sheet_name = f"{day[0]} {day[1]}"
        ws = wb.create_sheet(sheet_name)
        write_headers(ws)
        for row_n in range(2, 6):
            for col_n in range(1, 6):
                ws.cell(row=row_n, column=col_n).value = day[3][row_n-2][col_n-1]
        ws.cell(row=2, column=6).value = day[4]
        ws.cell(row=2, column=7).value = day[2]
        ws.cell(row=2, column=8).value = day[5]
    delete_first_sheet(wb)
    wb.save(file_name)
    time.sleep(5)
    os.startfile(os.path.abspath(file_name))
